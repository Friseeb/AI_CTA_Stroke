"""
copy_daylight_originals.py

Steps:
  1. Check which destination NIfTIs already exist and their sizes.
  2. Update Excel with NIfTI_status = "MISSING - needs re-download" for patients
     whose source file was not found in daylightbids\.
  3. Copy original NIfTIs (+ JSON sidecars) from daylightbids\ into SLAAOBIDS\.
     Skip variants (_Eq_1, _Tilt_1). Skip if destination already exists.
  4. Print final report.
"""

import shutil
from pathlib import Path

import pandas as pd
from tqdm import tqdm
from openpyxl import load_workbook

# ── Paths ──────────────────────────────────────────────────────────────────────
EXCEL_PATH  = Path(r"C:\Users\spost\Desktop\CT_image\SUB-ID_20260213a_FR.xlsx")
SRC_ORIG    = Path(r"C:\Users\spost\Desktop\CT_image\daylightbids")
SRC_DEFACED = Path(r"C:\Users\spost\Desktop\CT_image\daylightbids\derivatives\defaced")
DST_ROOT    = Path(r"C:\Users\spost\Desktop\CT_image\SLAAOBIDS")


# ── Helpers ────────────────────────────────────────────────────────────────────
def load_pairs() -> list[tuple[int, int]]:
    df = pd.read_excel(EXCEL_PATH)
    df.columns = [c.strip() for c in df.columns]
    df = df[df["DAYLIGHT"].notna() & df["BS-SLAAO"].notna()].copy()
    df["DAYLIGHT"] = df["DAYLIGHT"].astype(int)
    df["BS-SLAAO"] = df["BS-SLAAO"].astype(int)
    return list(zip(df["DAYLIGHT"], df["BS-SLAAO"]))


def mb(path: Path) -> float:
    return path.stat().st_size / (1024 ** 2)


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — Check existing destination NIfTIs
# ══════════════════════════════════════════════════════════════════════════════
def step1_check_existing(pairs: list[tuple[int, int]]) -> list[dict]:
    print("=" * 65)
    print("STEP 1 — Checking existing destination NIfTIs")
    print("=" * 65)

    corrupt = []
    present = []

    for day_id, slaao_id in pairs:
        dst = DST_ROOT / f"sub-{slaao_id}" / f"sub-{slaao_id}_acq-ecta_ct.nii.gz"
        if dst.exists():
            size = mb(dst)
            flag = " *** POTENTIALLY CORRUPT (<1 MB) ***" if size < 1.0 else ""
            present.append({"daylight": day_id, "slaao": slaao_id, "size_mb": size, "corrupt": size < 1.0})
            print(f"  EXISTS  sub-{slaao_id:>4}  (DAYLIGHT {day_id:>4})  {size:>8.1f} MB{flag}")
            if size < 1.0:
                corrupt.append((day_id, slaao_id, size))
        else:
            print(f"  absent  sub-{slaao_id:>4}  (DAYLIGHT {day_id:>4})")

    print()
    print(f"  Already present: {len(present)}  |  Absent: {len(pairs) - len(present)}")
    if corrupt:
        print(f"  *** {len(corrupt)} file(s) flagged as potentially corrupt (<1 MB) ***")
    print()
    return corrupt


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — Update Excel with NIfTI_status for missing DAYLIGHT sources
# ══════════════════════════════════════════════════════════════════════════════
def step2_update_excel(pairs: list[tuple[int, int]]) -> list[int]:
    print("=" * 65)
    print("STEP 2 — Updating Excel with NIfTI_status")
    print("=" * 65)

    # Determine which DAYLIGHT IDs have no source file
    missing_day_ids = []
    for day_id, _ in pairs:
        src = SRC_ORIG / f"sub-{day_id}_acq-CTA_ct.nii.gz"
        if not src.exists():
            missing_day_ids.append(day_id)

    print(f"  DAYLIGHT IDs with missing source NIfTI ({len(missing_day_ids)}): {missing_day_ids}")

    # Load workbook with openpyxl to preserve formatting
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    # Find or create the NIfTI_status column
    header_row = 1
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 2)]
    if "NIfTI_status" in headers:
        status_col = headers.index("NIfTI_status") + 1
        print(f"  Column 'NIfTI_status' already exists at column {status_col}.")
    else:
        status_col = ws.max_column + 1
        ws.cell(header_row, status_col).value = "NIfTI_status"
        print(f"  Created 'NIfTI_status' column at column {status_col}.")

    # Find the DAYLIGHT column index (column 1, header "DAYLIGHT")
    daylight_col = 1  # confirmed from earlier exploration

    updated = 0
    for row in range(2, ws.max_row + 1):
        cell_val = ws.cell(row, daylight_col).value
        if cell_val is None:
            continue
        try:
            row_day_id = int(cell_val)
        except (ValueError, TypeError):
            continue
        if row_day_id in missing_day_ids:
            ws.cell(row, status_col).value = "MISSING - needs re-download"
            updated += 1

    wb.save(EXCEL_PATH)
    print(f"  Wrote 'MISSING - needs re-download' on {updated} row(s). Excel saved.")
    print()
    return missing_day_ids


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 — Copy original NIfTIs + JSON sidecars
# ══════════════════════════════════════════════════════════════════════════════
def step3_copy(pairs: list[tuple[int, int]], missing_day_ids: list[int]):
    print("=" * 65)
    print("STEP 3 — Copying original NIfTIs and JSON sidecars")
    print("=" * 65)

    # Build copy plan
    ops = []
    for day_id, slaao_id in pairs:
        if day_id in missing_day_ids:
            continue  # source doesn't exist

        src_nii = SRC_ORIG / f"sub-{day_id}_acq-CTA_ct.nii.gz"
        src_json = SRC_ORIG / f"sub-{day_id}_acq-CTA_ct.json"
        dst_dir = DST_ROOT / f"sub-{slaao_id}"
        dst_nii = dst_dir / f"sub-{slaao_id}_acq-ecta_ct.nii.gz"
        dst_json = dst_dir / f"sub-{slaao_id}_acq-ecta_ct.json"

        ops.append(("nii", day_id, slaao_id, src_nii, dst_nii, dst_dir))
        if src_json.exists():
            ops.append(("json", day_id, slaao_id, src_json, dst_json, dst_dir))

    to_copy = [(kind, d, s, src, dst, d_dir)
               for kind, d, s, src, dst, d_dir in ops if not dst.exists()]
    skipped_exist = [(kind, d, s, src, dst, d_dir)
                     for kind, d, s, src, dst, d_dir in ops if dst.exists()]

    print(f"  Files to copy:          {len(to_copy)}")
    print(f"  Already exist (skip):   {len(skipped_exist)}")
    print()

    copied_nii = 0
    copied_json = 0
    for kind, day_id, slaao_id, src, dst, dst_dir in tqdm(
            to_copy, desc="Copying", unit="file", dynamic_ncols=True):
        dst_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy2(src, dst)
        if kind == "nii":
            copied_nii += 1
        else:
            copied_json += 1

    return copied_nii, copied_json, len(skipped_exist), missing_day_ids


# ══════════════════════════════════════════════════════════════════════════════
# STEP 4 — Final report
# ══════════════════════════════════════════════════════════════════════════════
def step4_report(pairs, copied_nii, copied_json, skipped_exist,
                 missing_day_ids, corrupt):
    print()
    print("=" * 65)
    print("STEP 4 — FINAL REPORT")
    print("=" * 65)
    print(f"  Original NIfTIs copied:       {copied_nii}")
    print(f"  JSON sidecars copied:         {copied_json}")
    print(f"  Skipped (dest. existed):      {skipped_exist}")
    print(f"  Skipped (source missing):     {len(missing_day_ids)}")
    print()

    # Patients that need defacing (original exists / was copied, but no defaced file)
    needs_defacing = []
    for day_id, slaao_id in pairs:
        if day_id in missing_day_ids:
            continue
        src_def = SRC_DEFACED / f"sub-{day_id}_acq-CTA_ct_defaced.nii.gz"
        if not src_def.exists():
            needs_defacing.append((day_id, slaao_id))

    print(f"  Patients needing defacing ({len(needs_defacing)}):")
    for day_id, slaao_id in needs_defacing:
        print(f"    DAYLIGHT {day_id:>4}  →  SLAAO sub-{slaao_id}")

    print()
    if corrupt:
        print(f"  *** Potentially corrupt files flagged in Step 1 ({len(corrupt)}) ***")
        for day_id, slaao_id, size in corrupt:
            print(f"    sub-{slaao_id}_acq-ecta_ct.nii.gz  ({size:.3f} MB)"
                  f"  [DAYLIGHT {day_id}]")
    else:
        print("  No potentially corrupt files detected in Step 1.")

    print()
    if missing_day_ids:
        print(f"  Missing source NIfTIs — Excel updated with 'MISSING - needs re-download':")
        for day_id in missing_day_ids:
            slaao_id = next(s for d, s in pairs if d == day_id)
            print(f"    DAYLIGHT {day_id:>4}  →  SLAAO sub-{slaao_id}")

    print()
    print("Done.")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    print(f"Loading pairs from: {EXCEL_PATH}")
    pairs = load_pairs()
    print(f"Found {len(pairs)} valid DAYLIGHT→SLAAO pairs.\n")

    corrupt          = step1_check_existing(pairs)
    missing_day_ids  = step2_update_excel(pairs)
    copied_nii, copied_json, skipped_exist, missing_day_ids = step3_copy(pairs, missing_day_ids)
    step4_report(pairs, copied_nii, copied_json, skipped_exist,
                 missing_day_ids, corrupt)


if __name__ == "__main__":
    main()
